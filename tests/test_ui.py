"""UI wiring, driven headlessly against a stub page.

Flet renders through Flutter to a canvas, so there is no DOM to assert on.
These tests instead exercise the real ExplorerApp handlers and check the two
things the UI layer is actually responsible for: what the status line says,
and how many rows reach the table.
"""

import asyncio

import pytest
from openpyxl import load_workbook

from hl7msg.model import FieldRow
from hl7msg.ui.app import MAX_DISPLAYED_ROWS, ExplorerApp


class StubPage:
    """The minimum Page surface ExplorerApp touches."""

    def __init__(self) -> None:
        self.services: list = []
        self.controls: list = []
        self.title = ""
        self.padding = 0
        self.updates = 0

    def add(self, *controls) -> None:
        self.controls.extend(controls)

    def update(self) -> None:
        self.updates += 1


@pytest.fixture
def app(sample_rows):
    application = ExplorerApp(StubPage())
    application.build()
    application.dataset.extend(sample_rows)
    return application


def drive(coro):
    return asyncio.run(coro)


def table_row_count(app) -> int:
    return len(app.results_area.controls[0].rows)


def test_build_registers_the_file_picker_as_a_service():
    application = ExplorerApp(StubPage())
    application.build()
    assert application.picker in application.page.services
    assert application.page.title == "HL7 Message Data Explorer"


def test_search_renders_matching_rows_and_reports_the_mode(app):
    app.search_field.value = "OBR.4.CE.2"
    drive(app.on_search(None))
    assert table_row_count(app) == 1
    assert "1 matches" in app.status.value
    assert "path spec" in app.status.value


def test_free_text_search_reports_value_mode(app):
    app.search_field.value = "Thyroid"
    drive(app.on_search(None))
    assert table_row_count(app) == 2
    assert "value text" in app.status.value


def test_bare_segment_reports_segment_mode(app):
    app.search_field.value = "OBR"
    drive(app.on_search(None))
    assert "segment name" in app.status.value
    assert table_row_count(app) > 0


def test_no_match_renders_an_empty_table_without_error(app):
    app.search_field.value = "zzz-nothing-matches-this"
    drive(app.on_search(None))
    assert table_row_count(app) == 0
    assert app.status.value.startswith("0 matches")


def test_searching_before_loading_says_so():
    application = ExplorerApp(StubPage())
    application.build()
    application.search_field.value = "OBX.3"
    drive(application.on_search(None))
    assert application.status.value == "No messages loaded yet."


def test_results_are_capped_and_the_true_total_is_reported(app):
    filler = [
        FieldRow(
            file_name=f"bulk_{index}.xml",
            path="OBX.3.CE.2",
            value=f"Condition {index}",
            full_path="ORU_R01.OBX.OBX.3.CE.2",
            numeric_path="OBX.3.2",
            depth=3,
        )
        for index in range(MAX_DISPLAYED_ROWS * 3)
    ]
    app.dataset.extend(filler)

    app.search_field.value = "OBX.3.CE.2"
    drive(app.on_search(None))

    # The fixture already contributes 3 rows at this path.
    expected_total = MAX_DISPLAYED_ROWS * 3 + 3
    assert table_row_count(app) == MAX_DISPLAYED_ROWS
    assert (
        f"showing {MAX_DISPLAYED_ROWS:,} of {expected_total:,} matches"
        in app.status.value
    )


def test_clear_empties_the_dataset_and_the_grid(app):
    app.search_field.value = "Thyroid"
    drive(app.on_search(None))
    assert table_row_count(app) > 0

    drive(app.on_clear(None))
    assert len(app.dataset) == 0
    assert app.search_field.value == ""
    assert table_row_count(app) == 0
    assert "Cleared" in app.status.value


# -- export --------------------------------------------------------------


class StubPicker:
    """Stands in for the native save dialog.

    Returns `destination` from save_file, or None to simulate a cancel, and
    records the arguments it was called with.
    """

    def __init__(self, destination=None):
        self.destination = destination
        self.calls = []

    async def save_file(self, **kwargs):
        self.calls.append(kwargs)
        return self.destination


def test_export_writes_every_match_not_just_the_displayed_page(app, tmp_path):
    """The whole point: the 500-row grid cap must not reach the workbook."""
    app.dataset.extend(
        [
            FieldRow(
                file_name=f"bulk_{n}.xml",
                path="OBX.3.CE.2",
                value=f"Condition {n}",
                full_path="ORU_R01.OBX.OBX.3.CE.2",
                numeric_path="OBX.3.2",
                depth=3,
            )
            for n in range(MAX_DISPLAYED_ROWS * 2)
        ]
    )
    app.search_field.value = "OBX.3.CE.2"
    drive(app.on_search(None))
    assert table_row_count(app) == MAX_DISPLAYED_ROWS

    destination = tmp_path / "out.xlsx"
    app.picker = StubPicker(str(destination))
    drive(app.on_export(None))

    expected = MAX_DISPLAYED_ROWS * 2 + 3  # fixture contributes 3
    written = load_workbook(destination)["Results"].max_row - 1
    assert written == expected
    assert f"Exported {expected:,} rows" in app.status.value


def test_export_passes_the_query_and_mode_through_to_the_workbook(app, tmp_path):
    app.search_field.value = "OBR.4.CE.2"
    drive(app.on_search(None))

    destination = tmp_path / "out.xlsx"
    app.picker = StubPicker(str(destination))
    drive(app.on_export(None))

    info = {
        r[0].value: r[1].value
        for r in load_workbook(destination)["Export Info"].iter_rows(max_row=5)
    }
    assert info["Query"] == "OBR.4.CE.2"
    assert info["Matched on"] == "path spec"


def test_export_suggests_a_filename_derived_from_the_query(app, tmp_path):
    app.search_field.value = "OBX.3"
    drive(app.on_search(None))

    picker = StubPicker(str(tmp_path / "out.xlsx"))
    app.picker = picker
    drive(app.on_export(None))

    suggested = picker.calls[0]["file_name"]
    assert suggested.startswith("hl7msg_OBX-3_")
    assert suggested.endswith(".xlsx")
    assert picker.calls[0]["allowed_extensions"] == ["xlsx"]


def test_cancelling_the_save_dialog_writes_nothing(app, tmp_path):
    app.search_field.value = "OBR.4.CE.2"
    drive(app.on_search(None))
    status_before = app.status.value

    app.picker = StubPicker(None)
    drive(app.on_export(None))

    assert list(tmp_path.iterdir()) == []
    assert app.status.value == status_before


def test_export_before_any_search_says_so():
    application = ExplorerApp(StubPage())
    application.build()
    application.picker = StubPicker("unused.xlsx")
    drive(application.on_export(None))
    assert "Nothing to export" in application.status.value


def test_export_of_an_empty_result_says_so(app, tmp_path):
    app.search_field.value = "zzz-nothing-matches-this"
    drive(app.on_search(None))

    app.picker = StubPicker(str(tmp_path / "out.xlsx"))
    drive(app.on_export(None))

    assert "Nothing to export" in app.status.value
    assert list(tmp_path.iterdir()) == []


def test_export_failure_is_reported_rather_than_raised(app, tmp_path):
    app.search_field.value = "OBR.4.CE.2"
    drive(app.on_search(None))

    # A directory that does not exist and cannot be created (a file is in the way).
    blocker = tmp_path / "blocker"
    blocker.write_text("not a directory")
    app.picker = StubPicker(str(blocker / "sub" / "out.xlsx"))

    drive(app.on_export(None))
    assert app.status.value.startswith("Export failed:")


def test_load_summary_counts_files_and_reports_failures():
    from hl7msg.store import LoadResult

    summary = ExplorerApp._describe_load(
        [
            LoadResult("a.xml", 40),
            LoadResult("b.xml", 38),
            LoadResult("notes.xml", 0, error="root element 'catalog' is not HL7"),
        ]
    )
    assert "Loaded 2 files, 78 rows." in summary
    assert "1 could not be read at all." in summary
    assert "See Load report." in summary


def test_load_summary_is_singular_for_one_file():
    from hl7msg.store import LoadResult

    assert "Loaded 1 file, 12 rows." in ExplorerApp._describe_load(
        [LoadResult("a.xml", 12)]
    )


def test_load_summary_says_recovered_files_are_only_partial():
    from hl7msg.store import LoadResult

    summary = ExplorerApp._describe_load(
        [
            LoadResult("a.xml", 400),
            LoadResult("b.xml", 60, recovered=True, issues=("line 9: expected '>'",)),
            LoadResult("c.xml", 25, recovered=True, issues=("line 3: expected '>'",)),
        ]
    )
    assert "Loaded 3 files, 485 rows." in summary
    assert "2 damaged and only partly recovered (85 rows)." in summary
    assert "could not be read" not in summary


def test_a_clean_load_says_nothing_about_reports():
    from hl7msg.store import LoadResult

    summary = ExplorerApp._describe_load([LoadResult("a.xml", 40)])
    assert "Load report" not in summary
    assert "recovered" not in summary


# -- load report ---------------------------------------------------------


class StubDialogPage(StubPage):
    """Records dialogs instead of showing them."""

    def __init__(self) -> None:
        super().__init__()
        self.dialogs: list = []

    def show_dialog(self, dialog) -> None:
        self.dialogs.append(dialog)

    def pop_dialog(self) -> None:
        self.dialogs.pop()


def load_app_with(results):
    app = ExplorerApp(StubDialogPage())
    app.build()
    app.last_load = results
    app.report_button.disabled = not any(r.needs_attention for r in results)
    return app


def test_report_button_is_disabled_until_something_goes_wrong():
    from hl7msg.store import LoadResult

    clean = load_app_with([LoadResult("a.xml", 40)])
    assert clean.report_button.disabled is True

    damaged = load_app_with([LoadResult("a.xml", 40, recovered=True)])
    assert damaged.report_button.disabled is False


def test_report_lists_failures_before_recoveries():
    from hl7msg.store import LoadResult

    app = load_app_with(
        [
            LoadResult("z_recovered.xml", 60, recovered=True, issues=("line 9: bad",)),
            LoadResult("a_failed.xml", 0, error="not HL7"),
            LoadResult("clean.xml", 400),
        ]
    )
    drive(app.on_show_report(None))

    dialog = app.page.dialogs[-1]
    table = dialog.content.content.controls[0]
    names = [r.cells[0].content.content.value for r in table.rows]
    outcomes = [r.cells[1].content.content.value for r in table.rows]

    # Clean files are not listed at all; failures come first.
    assert names == ["a_failed.xml", "z_recovered.xml"]
    assert outcomes == ["Failed", "Recovered"]


def test_report_shows_row_counts_and_the_first_problem():
    from hl7msg.store import LoadResult

    app = load_app_with(
        [
            LoadResult(
                "damaged.xml",
                1234,
                recovered=True,
                issues=("line 258, column 2: expected '>'", "line 9: something else"),
            )
        ]
    )
    drive(app.on_show_report(None))

    table = app.page.dialogs[-1].content.content.controls[0]
    cells = [c.content.content.value for c in table.rows[0].cells]
    assert cells[0] == "damaged.xml"
    assert cells[2] == "1,234"
    assert cells[3] == "line 258, column 2: expected '>'"


def test_report_on_a_clean_load_says_so_without_opening_a_dialog():
    from hl7msg.store import LoadResult

    app = load_app_with([LoadResult("a.xml", 40)])
    drive(app.on_show_report(None))
    assert app.page.dialogs == []
    assert "nothing to report" in app.status.value


def test_closing_the_report_pops_the_dialog():
    from hl7msg.store import LoadResult

    app = load_app_with([LoadResult("a.xml", 0, error="not HL7")])
    drive(app.on_show_report(None))
    assert len(app.page.dialogs) == 1

    drive(app.on_close_report(None))
    assert app.page.dialogs == []


# -- correlated view -----------------------------------------------------


CORR_SOURCE = b"""<ORU_R01 xmlns="urn:hl7-org:v2xml">
  <MSH><MSH.10>CERT-1</MSH.10></MSH>
  <ORU_R01.PATIENT_RESULT><ORU_R01.ORDER_OBSERVATION>
    <OBR><OBR.4><CE.2>Bloods</CE.2></OBR.4></OBR>
    <ORU_R01.OBSERVATION><OBX><OBX.3><CE.2>Thyroid</CE.2></OBX.3>
      <OBX.5>Yes</OBX.5><OBX.11>F</OBX.11></OBX></ORU_R01.OBSERVATION>
    <ORU_R01.OBSERVATION><OBX><OBX.3><CE.2>Headache</CE.2></OBX.3>
      <OBX.5>No</OBX.5><OBX.11>C</OBX.11></OBX></ORU_R01.OBSERVATION>
    <ORU_R01.OBSERVATION><OBX><OBX.3><CE.2>Cough</CE.2></OBX.3>
      <OBX.5>Yes, resolved</OBX.5><OBX.11>F</OBX.11></OBX></ORU_R01.OBSERVATION>
  </ORU_R01.ORDER_OBSERVATION></ORU_R01.PATIENT_RESULT>
</ORU_R01>"""


@pytest.fixture
def corr_app():
    from hl7msg.parser import parse_bytes

    application = ExplorerApp(StubDialogPage())
    application.build()
    application.dataset.extend(parse_bytes(CORR_SOURCE, "cert.xml").rows)
    return application


def corr_rows(app):
    return app.corr_results_area.controls[0].rows


def corr_values(app):
    return [
        tuple(c.content.content.value for c in row.cells) for row in corr_rows(app)
    ]


def test_building_a_table_projects_the_named_columns(corr_app):
    corr_app.columns_field.value = "OBR.4.CE.2 OBX.3.CE.2 OBX.5"
    drive(corr_app.on_build_table(None))

    assert corr_values(corr_app) == [
        ("Bloods", "Thyroid", "Yes"),
        ("Bloods", "Headache", "No"),
        ("Bloods", "Cough", "Yes, resolved"),
    ]
    assert "3 rows × 3 columns" in corr_app.corr_status.value


def test_a_filter_narrows_by_that_column_only(corr_app):
    corr_app.columns_field.value = "OBX.3.CE.2 OBX.5"
    drive(corr_app.on_build_table(None))

    corr_app.filter_field.value = "OBX.5: Ye"
    drive(corr_app.on_filter_changed(None))

    assert corr_values(corr_app) == [("Thyroid", "Yes"), ("Cough", "Yes, resolved")]
    assert "2 of 3 rows" in corr_app.corr_status.value


def test_filters_on_two_columns_intersect(corr_app):
    corr_app.columns_field.value = "OBX.3.CE.2 OBX.5 OBX.11"
    drive(corr_app.on_build_table(None))

    corr_app.filter_field.value = "OBX.5: Ye  OBX.11: F"
    drive(corr_app.on_filter_changed(None))

    assert [v[0] for v in corr_values(corr_app)] == ["Thyroid", "Cough"]

    corr_app.filter_field.value = "OBX.5: Ye  OBX.11: C"
    drive(corr_app.on_filter_changed(None))
    assert corr_values(corr_app) == []


def test_a_column_heading_filter_updates_the_search_box(corr_app):
    corr_app.columns_field.value = "OBX.3.CE.2 OBX.5"
    drive(corr_app.on_build_table(None))

    heading = corr_app.corr_results_area.controls[0].columns[1]
    box = heading.label.content.controls[1]
    assert box.data == "OBX.5"

    box.value = "Ye"
    drive(corr_app.on_column_filter_changed(_FakeEvent(box)))

    assert corr_app.filter_field.value == "OBX.5: Ye"
    assert len(corr_rows(corr_app)) == 2


def test_the_search_box_populates_the_column_headings(corr_app):
    corr_app.columns_field.value = "OBX.3.CE.2 OBX.5"
    drive(corr_app.on_build_table(None))

    corr_app.filter_field.value = "OBX.5: No"
    drive(corr_app.on_filter_changed(None))

    heading = corr_app.corr_results_area.controls[0].columns[1]
    assert heading.label.content.controls[1].value == "No"


def test_clearing_a_column_heading_removes_its_filter(corr_app):
    corr_app.columns_field.value = "OBX.3.CE.2 OBX.5"
    drive(corr_app.on_build_table(None))

    heading = corr_app.corr_results_area.controls[0].columns[1]
    box = heading.label.content.controls[1]
    box.value = "Ye"
    drive(corr_app.on_column_filter_changed(_FakeEvent(box)))
    assert len(corr_rows(corr_app)) == 2

    box = corr_app.corr_results_area.controls[0].columns[1].label.content.controls[1]
    box.value = ""
    drive(corr_app.on_column_filter_changed(_FakeEvent(box)))
    assert len(corr_rows(corr_app)) == 3
    assert corr_app.filter_field.value == ""


def test_rebuilding_drops_filters_for_columns_no_longer_shown(corr_app):
    corr_app.columns_field.value = "OBX.3.CE.2 OBX.5"
    drive(corr_app.on_build_table(None))
    corr_app.filter_field.value = "OBX.5: Ye"
    drive(corr_app.on_filter_changed(None))
    assert corr_app.column_filters == {"OBX.5": "Ye"}

    corr_app.columns_field.value = "OBX.3.CE.2 OBX.11"
    drive(corr_app.on_build_table(None))
    assert corr_app.column_filters == {}
    assert len(corr_rows(corr_app)) == 3


def test_building_without_messages_says_so():
    application = ExplorerApp(StubDialogPage())
    application.build()
    application.columns_field.value = "OBX.5"
    drive(application.on_build_table(None))
    assert application.corr_status.value == "No messages loaded yet."


def test_building_with_no_columns_says_so(corr_app):
    corr_app.columns_field.value = "   "
    drive(corr_app.on_build_table(None))
    assert "at least one HL7 path" in corr_app.corr_status.value


def test_a_mistyped_column_is_reported(corr_app):
    corr_app.columns_field.value = "OBX.3.CE.2 OBX.99"
    drive(corr_app.on_build_table(None))
    assert "No values found for OBX.99" in corr_app.corr_status.value


def test_a_filter_naming_an_unprojected_column_is_reported(corr_app):
    corr_app.columns_field.value = "OBX.3.CE.2"
    drive(corr_app.on_build_table(None))

    corr_app.filter_field.value = "OBX.5: Ye"
    drive(corr_app.on_filter_changed(None))
    assert "Not a projected column: OBX.5" in corr_app.corr_status.value


def test_correlated_export_writes_the_filtered_table(corr_app, tmp_path):
    from openpyxl import load_workbook

    corr_app.columns_field.value = "OBX.3.CE.2 OBX.5"
    drive(corr_app.on_build_table(None))
    corr_app.filter_field.value = "OBX.5: Ye"
    drive(corr_app.on_filter_changed(None))

    destination = tmp_path / "corr.xlsx"
    corr_app.picker = StubPicker(str(destination))
    drive(corr_app.on_export_correlated(None))

    sheet = load_workbook(destination)["Results"]
    written = [[c.value for c in row] for row in sheet.iter_rows(min_row=2)]
    assert written == [
        ["cert.xml", "Thyroid", "Yes"],
        ["cert.xml", "Cough", "Yes, resolved"],
    ]
    assert "Exported 2 rows" in corr_app.corr_status.value


def test_correlated_export_before_building_says_so(corr_app, tmp_path):
    corr_app.picker = StubPicker(str(tmp_path / "x.xlsx"))
    drive(corr_app.on_export_correlated(None))
    assert "Nothing to export" in corr_app.corr_status.value
    assert list(tmp_path.iterdir()) == []


def test_clear_resets_the_correlated_view(corr_app):
    corr_app.columns_field.value = "OBX.3.CE.2 OBX.5"
    drive(corr_app.on_build_table(None))
    corr_app.filter_field.value = "OBX.5: Ye"
    drive(corr_app.on_filter_changed(None))

    drive(corr_app.on_clear(None))
    assert corr_app.correlated_table is None
    assert corr_app.column_filters == {}
    assert corr_app.columns_field.value == ""
    assert corr_app.filter_field.value == ""


class _FakeEvent:
    """Stands in for a Flet control event."""

    def __init__(self, control):
        self.control = control


def test_clear_disables_the_report_button(app):
    from hl7msg.store import LoadResult

    app.last_load = [LoadResult("a.xml", 0, error="not HL7")]
    app.report_button.disabled = False

    drive(app.on_clear(None))
    assert app.last_load == []
    assert app.report_button.disabled is True
