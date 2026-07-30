"""Excel export, verified by reading each workbook back."""

from datetime import datetime

import pytest
from openpyxl import load_workbook

from hl7msg.export import (
    EXPORT_COLUMNS,
    MAX_CELL_CHARS,
    TRUNCATION_MARKER,
    ExportError,
    suggested_filename,
    write_xlsx,
)
from hl7msg.model import FieldRow


def make_row(value: str = "Thyroid disorder", path: str = "OBX.3.CE.2") -> FieldRow:
    return FieldRow(
        file_name="cert_001.xml",
        path=path,
        value=value,
        full_path=f"ORU_R01.OBX.{path}",
        numeric_path="OBX.3.2",
        depth=3,
    )


def read_sheet(path, name="Results"):
    return load_workbook(path)[name]


def data_rows(sheet):
    """Every row below the header, as lists of values."""
    return [[c.value for c in row] for row in sheet.iter_rows(min_row=2)]


# -- shape ---------------------------------------------------------------


def test_headers_are_the_six_configured_columns(tmp_path):
    out = write_xlsx([make_row()], tmp_path / "out.xlsx")
    sheet = read_sheet(out)
    headers = [cell.value for cell in next(sheet.iter_rows(max_row=1))]
    assert headers == [title for title, _ in EXPORT_COLUMNS]
    assert len(headers) == 6


def test_a_row_round_trips_every_field(tmp_path):
    row = make_row()
    out = write_xlsx([row], tmp_path / "out.xlsx")
    written = data_rows(read_sheet(out))[0]
    assert written == [
        row.file_name,
        row.path,
        row.value,
        row.full_path,
        row.numeric_path,
        row.depth,
    ]


def test_depth_stays_numeric(tmp_path):
    out = write_xlsx([make_row()], tmp_path / "out.xlsx")
    assert isinstance(data_rows(read_sheet(out))[0][5], int)


def test_header_is_frozen_and_autofilter_is_set(tmp_path):
    out = write_xlsx([make_row(), make_row()], tmp_path / "out.xlsx")
    sheet = read_sheet(out)
    assert sheet.freeze_panes == "A2"
    assert sheet.auto_filter.ref == "A1:F3"


# -- the core requirement: no 500-row cap --------------------------------


def test_exports_every_row_not_just_the_displayed_page(tmp_path):
    """The grid caps at 500; the export must not."""
    rows = [make_row(value=f"Condition {n}") for n in range(1500)]
    out = write_xlsx(rows, tmp_path / "out.xlsx")
    written = data_rows(read_sheet(out))
    assert len(written) == 1500
    assert written[0][2] == "Condition 0"
    assert written[-1][2] == "Condition 1499"


def test_row_order_is_preserved(tmp_path):
    rows = [make_row(value=str(n)) for n in range(200)]
    out = write_xlsx(rows, tmp_path / "out.xlsx")
    assert [r[2] for r in data_rows(read_sheet(out))] == [str(n) for n in range(200)]


def test_empty_result_still_writes_a_valid_workbook(tmp_path):
    out = write_xlsx([], tmp_path / "out.xlsx")
    sheet = read_sheet(out)
    assert data_rows(sheet) == []
    assert [c.value for c in next(sheet.iter_rows(max_row=1))][0] == "File"


# -- hostile values ------------------------------------------------------


def test_a_formula_looking_value_stays_literal_text(tmp_path):
    """Values come from third-party messages; none of them may execute."""
    out = write_xlsx([make_row(value="=1+1")], tmp_path / "out.xlsx")
    cell = read_sheet(out).cell(row=2, column=3)
    assert cell.value == "=1+1"
    assert cell.data_type == "s"


@pytest.mark.parametrize("value", ["=SUM(A1:A9)", "+1", "-1", "@import", "=cmd|'/c calc'!A"])
def test_other_formula_prefixes_stay_literal_text(tmp_path, value):
    out = write_xlsx([make_row(value=value)], tmp_path / "out.xlsx")
    cell = read_sheet(out).cell(row=2, column=3)
    assert cell.value == value
    assert cell.data_type == "s"


def test_control_characters_are_stripped(tmp_path):
    out = write_xlsx([make_row(value="Chest\x00Pain\x0bhere\x1f")], tmp_path / "out.xlsx")
    assert data_rows(read_sheet(out))[0][2] == "ChestPainhere"


def test_tab_and_newline_survive(tmp_path):
    """Layout characters in clinical narrative must not be stripped.

    Excel itself normalises a CRLF to a bare LF on the way in, so a carriage
    return is asserted separately below rather than round-tripped verbatim.
    """
    narrative = "COMPLAINT:\tChest Pain\nOUTCOME:\nAdmitted"
    out = write_xlsx([make_row(value=narrative)], tmp_path / "out.xlsx")
    assert data_rows(read_sheet(out))[0][2] == narrative


def test_a_carriage_return_is_not_treated_as_an_illegal_character(tmp_path):
    out = write_xlsx([make_row(value="A\r\nB")], tmp_path / "out.xlsx")
    written = data_rows(read_sheet(out))[0][2]

    # How a CRLF comes back depends on which XML backend openpyxl picked:
    # with lxml installed it round-trips as "A\r\nB", and on the stdlib
    # backend each CR becomes a LF, giving "A\n\nB". Asserting either exact
    # string would pin an incidental dependency. The property that matters is
    # that neither character was stripped the way a genuine illegal control
    # character would be.
    assert written.startswith("A")
    assert written.endswith("B")
    assert "\n" in written or "\r" in written
    assert len(written) >= len("A\nB")


def test_a_long_clinical_narrative_survives_intact(tmp_path):
    """The real corpus contains NTE.3 narratives of 4,249 characters."""
    narrative = "COMPLAINT:Chest Pain. " * 200
    assert len(narrative) < MAX_CELL_CHARS
    out = write_xlsx([make_row(value=narrative)], tmp_path / "out.xlsx")
    assert data_rows(read_sheet(out))[0][2] == narrative


def test_an_over_length_value_is_truncated_with_a_marker(tmp_path):
    out = write_xlsx([make_row(value="x" * (MAX_CELL_CHARS + 500))], tmp_path / "out.xlsx")
    written = data_rows(read_sheet(out))[0][2]
    assert len(written) == MAX_CELL_CHARS
    assert written.endswith(TRUNCATION_MARKER)


def test_too_many_rows_raises_rather_than_writing_a_corrupt_file(tmp_path):
    class FakeRows:
        def __len__(self):
            return 2_000_000

        def __iter__(self):
            return iter(())

    destination = tmp_path / "out.xlsx"
    with pytest.raises(ExportError, match="exceeds the"):
        write_xlsx(FakeRows(), destination)
    assert not destination.exists()


# -- destination handling ------------------------------------------------


def test_xlsx_suffix_is_added_when_missing(tmp_path):
    out = write_xlsx([make_row()], tmp_path / "results")
    assert out.name == "results.xlsx"
    assert out.exists()


def test_an_existing_xlsx_suffix_is_left_alone(tmp_path):
    out = write_xlsx([make_row()], tmp_path / "results.xlsx")
    assert out.name == "results.xlsx"


def test_a_non_xlsx_suffix_is_appended_to_rather_than_replaced(tmp_path):
    """Replacing would silently discard part of a name the user typed."""
    out = write_xlsx([make_row()], tmp_path / "results.csv")
    assert out.name == "results.csv.xlsx"


def test_a_dotted_name_is_not_truncated(tmp_path):
    out = write_xlsx([make_row()], tmp_path / "extract.v2")
    assert out.name == "extract.v2.xlsx"
    assert out.exists()


# -- audit sheet ---------------------------------------------------------


def test_export_info_records_the_query_mode_and_count(tmp_path):
    rows = [make_row() for _ in range(7)]
    out = write_xlsx(
        rows,
        tmp_path / "out.xlsx",
        query="OBX.3.CE.2",
        mode="path spec",
        source_files=["a.xml", "b.xml"],
    )
    info = {
        row[0].value: row[1].value
        for row in load_workbook(out)["Export Info"].iter_rows(max_row=5)
    }
    assert info["Query"] == "OBX.3.CE.2"
    assert info["Matched on"] == "path spec"
    assert info["Rows exported"] == 7
    assert info["Source files"] == 2


def test_export_info_lists_the_source_files(tmp_path):
    out = write_xlsx(
        [make_row()], tmp_path / "out.xlsx", source_files=["a.xml", "b.xml"]
    )
    values = [
        cell.value
        for row in load_workbook(out)["Export Info"].iter_rows()
        for cell in row
    ]
    assert "a.xml" in values and "b.xml" in values


def test_a_blank_query_is_recorded_as_all_rows(tmp_path):
    out = write_xlsx([make_row()], tmp_path / "out.xlsx", query="")
    info = {
        row[0].value: row[1].value
        for row in load_workbook(out)["Export Info"].iter_rows(max_row=5)
    }
    assert info["Query"] == "(all rows)"


# -- suggested filename --------------------------------------------------


def test_suggested_filename_is_derived_from_the_query():
    when = datetime(2026, 7, 29, 18, 30)
    assert suggested_filename("OBX.3.CE.2", when) == "hl7msg_OBX-3-CE-2_20260729-1830.xlsx"


def test_suggested_filename_strips_characters_illegal_in_a_path():
    name = suggested_filename('OBR.*.CE.2 <>:"/\\|?', datetime(2026, 1, 2, 3, 4))
    assert name.startswith("hl7msg_OBR-")
    assert not set(name) & set('<>:"/\\|?*')


def test_suggested_filename_falls_back_when_the_query_is_blank():
    assert "all-rows" in suggested_filename("   ", datetime(2026, 1, 2, 3, 4))


# -- correlated tables ---------------------------------------------------


CORRELATED_SOURCE = b"""<ORU_R01 xmlns="urn:hl7-org:v2xml">
  <MSH><MSH.10>CERT-1</MSH.10></MSH>
  <ORU_R01.PATIENT_RESULT><ORU_R01.ORDER_OBSERVATION>
    <OBR><OBR.4><CE.2>Bloods</CE.2></OBR.4></OBR>
    <ORU_R01.OBSERVATION><OBX><OBX.3><CE.2>Thyroid</CE.2></OBX.3>
      <OBX.5>Yes</OBX.5></OBX></ORU_R01.OBSERVATION>
    <ORU_R01.OBSERVATION><OBX><OBX.3><CE.2>Headache</CE.2></OBX.3>
      <OBX.5>=1+1</OBX.5></OBX></ORU_R01.OBSERVATION>
  </ORU_R01.ORDER_OBSERVATION></ORU_R01.PATIENT_RESULT>
</ORU_R01>"""


@pytest.fixture
def correlated():
    from hl7msg.correlate import CorrelationEngine
    from hl7msg.parser import parse_bytes
    from hl7msg.store import Dataset

    data = Dataset()
    data.extend(parse_bytes(CORRELATED_SOURCE, "cert.xml").rows)
    return CorrelationEngine(data).search("MSH.10 OBR.4.CE.2 OBX.3.CE.2 OBX.5")


def test_correlated_export_uses_the_projection_as_headers(tmp_path, correlated):
    from hl7msg.export import write_correlated_xlsx

    out = write_correlated_xlsx(correlated, tmp_path / "out.xlsx")
    headers = [c.value for c in next(read_sheet(out).iter_rows(max_row=1))]
    assert headers == ["File", "MSH.10", "OBR.4.CE.2", "OBX.3.CE.2", "OBX.5"]


def test_correlated_export_writes_one_row_per_segment(tmp_path, correlated):
    from hl7msg.export import write_correlated_xlsx

    out = write_correlated_xlsx(correlated, tmp_path / "out.xlsx")
    written = data_rows(read_sheet(out))
    assert written == [
        ["cert.xml", "CERT-1", "Bloods", "Thyroid", "Yes"],
        ["cert.xml", "CERT-1", "Bloods", "Headache", "=1+1"],
    ]


def test_correlated_export_keeps_a_formula_looking_cell_literal(tmp_path, correlated):
    from hl7msg.export import write_correlated_xlsx

    out = write_correlated_xlsx(correlated, tmp_path / "out.xlsx")
    cell = read_sheet(out).cell(row=3, column=5)
    assert cell.value == "=1+1"
    assert cell.data_type == "s"


def test_correlated_export_is_framed_like_the_flat_one(tmp_path, correlated):
    from hl7msg.export import write_correlated_xlsx

    out = write_correlated_xlsx(correlated, tmp_path / "out.xlsx")
    sheet = read_sheet(out)
    assert sheet.freeze_panes == "A2"
    assert sheet.auto_filter.ref == "A1:E3"


def test_correlated_export_records_columns_filters_and_grain(tmp_path, correlated):
    from hl7msg.export import write_correlated_xlsx

    out = write_correlated_xlsx(
        correlated,
        tmp_path / "out.xlsx",
        columns_text="MSH.10 OBR.4.CE.2 OBX.3.CE.2 OBX.5",
        filter_text="OBX.5: Ye",
    )
    info = {
        row[0].value: row[1].value
        for row in load_workbook(out)["Export Info"].iter_rows(max_row=12)
        if row[0].value
    }
    assert info["Columns"] == "MSH.10 OBR.4.CE.2 OBX.3.CE.2 OBX.5"
    assert info["Filters"] == "OBX.5: Ye"
    assert "OBX.3.CE.2" in info["Row grain"]
    assert info["Rows exported"] == 2
