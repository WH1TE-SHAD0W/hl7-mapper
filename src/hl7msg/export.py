"""Export a filtered result set to an Excel workbook.

Pure Python plus openpyxl -- no flet -- so exporting is unit-testable without
a GUI, like the parser and search engine.

What lands in the workbook is the *whole* result set, not the capped subset
the results grid displays. The cap exists only because a table widget cannot
render 58,000 rows interactively; it is not a limit on the data.
"""

import re
from datetime import datetime
from pathlib import Path
from typing import Sequence

from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from .model import FieldRow

#: Worksheet columns, as (heading, FieldRow attribute).
EXPORT_COLUMNS: list[tuple[str, str]] = [
    ("File", "file_name"),
    ("Path", "path"),
    ("Value", "value"),
    ("Full Path", "full_path"),
    ("Numeric Path", "numeric_path"),
    ("Depth", "depth"),
]

#: Hard limits of the .xlsx format itself.
MAX_EXCEL_ROWS = 1_048_576
MAX_CELL_CHARS = 32_767

TRUNCATION_MARKER = " …[truncated]"

_COLUMN_WIDTHS = [28, 24, 70, 40, 20, 8]

#: Characters Excel refuses to accept in a cell. Tab, newline and carriage
#: return are legal and are deliberately preserved -- clinical narrative
#: reconstructed from HL7 messages contains them.
_ILLEGAL = ILLEGAL_CHARACTERS_RE

#: Dots are excluded deliberately as well as the characters Windows forbids:
#: a name like ``hl7msg_OBX.3.CE.2_...`` reads as though it has several
#: extensions, and any later suffix handling would have to guess where the
#: real one starts.
_UNSAFE_FILENAME_CHARS = re.compile(r"[^A-Za-z0-9_-]+")


class ExportError(RuntimeError):
    """Raised when a result set cannot be written as a valid workbook."""


def _sanitise(text: str) -> str:
    """Make a value safe to place in a cell.

    Strips control characters openpyxl would reject, and truncates anything
    past Excel's per-cell limit with a visible marker rather than letting the
    write fail or silently produce a corrupt file.
    """
    text = _ILLEGAL.sub("", text)
    if len(text) > MAX_CELL_CHARS:
        keep = MAX_CELL_CHARS - len(TRUNCATION_MARKER)
        text = text[:keep] + TRUNCATION_MARKER
    return text


def _text_cell(sheet, text: str) -> WriteOnlyCell:
    """A cell that is always literal text.

    openpyxl infers a formula from a leading ``=``, so a field value of
    ``=1+1`` would otherwise be written as a live formula. These values come
    from third-party HL7 messages, so the data type is pinned to string.
    """
    cell = WriteOnlyCell(sheet, value=_sanitise(text))
    cell.data_type = "s"
    return cell


def suggested_filename(query: str, when: datetime | None = None) -> str:
    """A default save name derived from the query and the current time.

    ``OBX.3.CE.2`` becomes ``hl7msg_OBX-3-CE-2_20260729-1830.xlsx``.
    """
    when = when or datetime.now()
    stem = _UNSAFE_FILENAME_CHARS.sub("-", query.strip()).strip("-_")
    stem = stem[:60] or "all-rows"
    return f"hl7msg_{stem}_{when:%Y%m%d-%H%M}.xlsx"


def _open_sheet(
    workbook: Workbook, headers: Sequence[str], widths: Sequence[int], row_count: int
):
    """Create a results sheet with its header written and framing applied.

    In write-only mode the sheet header XML is emitted on the first append, so
    freeze panes, filters and widths must all be configured before any row is
    written -- set afterwards they are silently discarded.
    """
    sheet = workbook.create_sheet("Results")

    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    # Freeze the header and switch AutoFilter on: filtering is the analyst's
    # very next action after opening the file.
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{row_count + 1}"

    header_cells = []
    for title in headers:
        cell = WriteOnlyCell(sheet, value=title)
        cell.font = Font(bold=True)
        cell.data_type = "s"
        header_cells.append(cell)
    sheet.append(header_cells)
    return sheet


def _write_results_sheet(workbook: Workbook, rows: Sequence[FieldRow]) -> None:
    headers = [title for title, _ in EXPORT_COLUMNS]
    sheet = _open_sheet(workbook, headers, _COLUMN_WIDTHS, len(rows))

    for row in rows:
        cells = []
        for _title, attribute in EXPORT_COLUMNS:
            value = getattr(row, attribute)
            # Depth stays numeric so it can be sorted and filtered as a number.
            if isinstance(value, int) and not isinstance(value, bool):
                cells.append(value)
            else:
                cells.append(_text_cell(sheet, str(value)))
        sheet.append(cells)


def _write_info_sheet(
    workbook: Workbook,
    query: str,
    mode: str,
    row_count: int,
    source_files: Sequence[str],
    recovered_files: Sequence[str] = (),
    extra: Sequence[tuple[str, object]] = (),
) -> None:
    """Record how this extract was produced, for auditability."""
    sheet = workbook.create_sheet("Export Info")
    sheet.column_dimensions["A"].width = 20
    sheet.column_dimensions["B"].width = 70

    def labelled(label: str, value) -> None:
        left = WriteOnlyCell(sheet, value=label)
        left.font = Font(bold=True)
        left.data_type = "s"
        if isinstance(value, int) and not isinstance(value, bool):
            sheet.append([left, value])
        else:
            sheet.append([left, _text_cell(sheet, str(value))])

    labelled("Exported", f"{datetime.now():%Y-%m-%d %H:%M:%S}")
    labelled("Query", query or "(all rows)")
    labelled("Matched on", mode or "n/a")
    labelled("Rows exported", row_count)
    labelled("Source files", len(source_files))
    for label, value in extra:
        labelled(label, value)

    # Rows from a recovered file are only what survived the damage. Recording
    # it here means an extract can still be traced back to whether partial
    # data fed it, long after the app has been closed.
    contributing = [name for name in recovered_files if name in set(source_files)]
    labelled("Partially recovered sources", len(contributing))
    if contributing:
        labelled("Recovered file names", ", ".join(sorted(contributing)))

    if source_files:
        sheet.append([])
        heading = WriteOnlyCell(sheet, value="Source file")
        heading.font = Font(bold=True)
        heading.data_type = "s"
        sheet.append([heading])
        for name in source_files:
            sheet.append([_text_cell(sheet, name)])


def _prepare_destination(destination: str | Path, row_count: int) -> Path:
    """Normalise the path and refuse a result set no worksheet can hold."""
    destination = Path(destination)
    if destination.suffix.lower() != ".xlsx":
        # Append rather than replace. Path.with_suffix would cut at the last
        # dot, turning a legitimate name like "extract.v2" into "extract.xlsx"
        # and losing part of what the user typed.
        destination = destination.with_name(destination.name + ".xlsx")

    if row_count + 1 > MAX_EXCEL_ROWS:
        raise ExportError(
            f"{row_count:,} rows exceeds the {MAX_EXCEL_ROWS:,}-row limit of an "
            f"Excel worksheet; narrow the search before exporting"
        )

    # Make sure the destination is usable before building anything: there is
    # no point serialising hundreds of thousands of rows only to discover the
    # directory cannot be created.
    destination.parent.mkdir(parents=True, exist_ok=True)
    return destination


def write_correlated_xlsx(
    table,
    destination: str | Path,
    *,
    columns_text: str = "",
    filter_text: str = "",
    recovered_files: Sequence[str] = (),
) -> Path:
    """Write a CorrelatedTable, one column per projected HL7 path.

    A ``File`` column is prepended so a row can always be traced back to the
    message it came from; the projected columns follow in the order they were
    typed.
    """
    destination = _prepare_destination(destination, len(table.rows))

    headers = ["File", *table.columns]
    widths = [30, *(34 for _ in table.columns)]

    workbook = Workbook(write_only=True)
    try:
        sheet = _open_sheet(workbook, headers, widths, len(table.rows))
        for row in table.rows:
            sheet.append(
                [
                    _text_cell(sheet, row.file_name),
                    *(_text_cell(sheet, value) for value in row.cells),
                ]
            )

        source_files = sorted({row.file_name for row in table.rows})
        _write_info_sheet(
            workbook,
            query=columns_text,
            mode="correlated projection",
            row_count=len(table.rows),
            source_files=source_files,
            recovered_files=recovered_files,
            extra=(
                ("Columns", " ".join(table.columns)),
                ("Filters", filter_text or "(none)"),
                ("Row grain", ", ".join(table.anchors)),
                ("Ambiguous cells", table.ambiguous_cells),
                ("Rows from partial messages", table.recovered_rows),
            ),
        )
        workbook.save(destination)
    finally:
        workbook.close()
    return destination


def write_xlsx(
    rows: Sequence[FieldRow],
    destination: str | Path,
    *,
    query: str = "",
    mode: str = "",
    source_files: Sequence[str] = (),
    recovered_files: Sequence[str] = (),
) -> Path:
    """Write ``rows`` to an .xlsx workbook at ``destination``.

    Uses openpyxl's write-only mode, which streams rows out instead of
    building a cell object for every cell -- a full-dataset export here is
    393,000 rows, which the normal object model would turn into millions of
    live objects.

    Returns the path written. Raises ExportError if the result set cannot fit
    a worksheet.
    """
    destination = _prepare_destination(destination, len(rows))

    workbook = Workbook(write_only=True)
    try:
        _write_results_sheet(workbook, rows)
        _write_info_sheet(
            workbook, query, mode, len(rows), source_files, recovered_files
        )
        workbook.save(destination)
    finally:
        # A write-only workbook holds open temporary streams. Without this
        # they are only released at collection, which surfaces as an
        # unraisable exception whenever a save fails part way through.
        workbook.close()
    return destination
